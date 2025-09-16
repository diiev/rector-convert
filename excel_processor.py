import openpyxl 
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font 
from openpyxl.utils import get_column_letter
import copy
from utils import format_fio


class ExcelProcessor:
    def __init__(self, gui_callback=None):
        self.gui_callback = gui_callback
        # Предопределенные стили для избежания повторного создания
        self.border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.header_font = Font(size=11, bold=True, name='Arial')
        self.title_font = Font(size=18, bold=True)
        self.words_to_remove = ['(лекция)', '(практика)', ', вид занятия']
        self.day_colors = {
            'Пн': "D9E1F2",
            'Вт': "FCE4D6", 
            'Ср': "FFF2CC",
            'Чт': "E2EFDA",
            'Пт': "D6DCE4",
            'Сб': "FFB3B3"
        }
    
    def update_gui_text(self, text):
        if self.gui_callback:
            self.gui_callback(text)
    
    def _copy_sheet_data(self, source_sheet, target_sheet, max_row, max_column):
        """Копирование данных между листами"""
        for row in range(2, max_row + 1):
            for column in range(1, max_column + 1):
                cell = source_sheet.cell(row - 1, column)
                new_cell = target_sheet.cell(row, column, value=cell.value)
                new_cell.font = copy.copy(cell.font)
                target_sheet.column_dimensions[get_column_letter(column)].width = \
                    source_sheet.column_dimensions[get_column_letter(column)].width
    
    def create_sheets_for_teacher(self, file_path):
        """Создание отдельных листов для каждого преподавателя"""
        workbook = openpyxl.load_workbook(file_path)
        new_workbook = openpyxl.Workbook()
        total_sheets = len(workbook.sheetnames)
        
        for count_sheets, sheet_name in enumerate(workbook.sheetnames, 1):
            sheet = workbook[sheet_name]
            new_sheet = new_workbook.create_sheet(sheet_name)
            new_sheet.bestFit = True
            
            self.update_gui_text(f"Копируем значения... {round((count_sheets/total_sheets) * 100,1)}%") 
            
            for row in range(1, sheet.max_row + 1):
                cell_value = sheet.cell(row, 1).value
                if (cell_value and isinstance(cell_value, str) and 'Преподаватель' in cell_value):
                    fio = cell_value.replace('Преподаватель -', '').strip()
                    formatted_fio = format_fio(fio)
                    new_sheet = new_workbook.create_sheet(formatted_fio)
                else:
                    self._copy_sheet_data(sheet, new_sheet, sheet.max_row, sheet.max_column)

        # Удаление временных листов
        for sheet_name in ['1', 'Sheet']:
            if sheet_name in new_workbook.sheetnames:
                new_workbook.remove(new_workbook[sheet_name])
        
        save_file = file_path.replace('.xlsx', '_updated.xlsx')
        new_workbook.save(save_file)
        new_workbook.close()
        return save_file

    def _setup_teacher_sheet_formatting(self, sheet):
        """Настройка форматирования для листа преподавателя"""
        sheet.merge_cells('A1:M1')
        sheet.merge_cells('A2:M2')
        
        # Объединение ячеек и установка заголовков
        merge_ranges = ['A3:A4', 'C3:C4', 'E3:E4', 'G3:G4', 'I3:I4', 'K3:K4', 'M3:M4']
        for range_str in merge_ranges:
            sheet.merge_cells(range_str)
        
        # Установка заголовков "Ауд"
        for col in [3, 5, 7, 9, 11, 13]:
            sheet.cell(3, col).value = 'Ауд'
            sheet.cell(3, col).font = self.header_font

    def remove_empty_rows(self, file_path):
        """Удаление пустых строк и форматирование"""
        workbook = openpyxl.load_workbook(file_path)
        total_sheets = len(workbook.sheetnames)
        
        for count_sheets, sheet_name in enumerate(workbook.sheetnames, 1):
            sheet = workbook[sheet_name]
            self.update_gui_text(f"Обработка листа {sheet_name} {round((count_sheets/total_sheets) * 100,1)}%") 
            
            self._setup_teacher_sheet_formatting(sheet)
            
            # Удаление полностью пустых строк
            for row in reversed(range(1, sheet.max_row + 1)):
                if all(cell.value in (None, '', '-') for cell in sheet[row]):
                    sheet.delete_rows(row)
            
            # Применение стилей
            for row in range(3, sheet.max_row + 1):
                for column in range(1, sheet.max_column + 1):
                    sheet.cell(row, column).border = self.border_style
                    sheet.cell(row, column).alignment = Alignment(
                        wrap_text=True, horizontal='center', vertical='center'
                    )
        
        workbook.save(file_path)
        workbook.close()

    def _setup_group_sheet_formatting(self, sheet):
        """Настройка форматирования для листа группы"""
        if sheet.max_column > 4:
            sheet.merge_cells('A1:F1')
            sheet.merge_cells('A3:F3')
            column_widths = {'D': 8.5, 'F': 8.5, 'B': 5.5, 'A': 7}
        else:
            sheet.merge_cells('A1:D1')
            sheet.merge_cells('A3:D3')
            column_widths = {'B': 10, 'A': 13, 'C': 50.5, 'D': 15.5}
        
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width
        
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sheet['A1'].font = self.title_font
        
        # Скрытие служебных строк
        for row in [2, 3, 4]:
            sheet.row_dimensions[row].hidden = True

    def _process_cell_content(self, cell, remove_default_words, custom_words):
        """Обработка содержимого ячейки"""
        if not cell.value:
            return
        
        cell_value = str(cell.value)
        words_to_remove = self.words_to_remove if remove_default_words else custom_words.split("-")
        
        for word in words_to_remove:
            if word.lower() in cell_value.lower():
                cell_value = cell_value.replace(word, "")
        
        cell.value = cell_value

    def _should_hide_row(self, row_data, max_column):
        """Определить, нужно ли скрыть строку"""
        if max_column > 4:
            return (row_data[2].value in (None, '') and row_data[4].value in (None, ''))
        else:
            return (row_data[2].value in (None, '') and row_data[3].value in (None, ''))

    def remove_empty_cells_and_words(self, file_path, remove_default_words, set_colors, word_remove_entry):
        """Удаление пустых ячеек и слов из расписания групп"""
        workbook = openpyxl.load_workbook(file_path) 
        total_sheets = len(workbook.sheetnames)
        
        for count_sheets, sheet_name in enumerate(workbook.sheetnames, 1):
            sheet = workbook[sheet_name]
            self.update_gui_text(f"Обработка листа {sheet_name} {round((count_sheets/total_sheets) * 100,1)}%") 
            
            self._setup_group_sheet_formatting(sheet)
            
            for row in reversed(range(5, sheet.max_row + 1)): 
                row_data = sheet[row]
                
                # Обработка ячеек
                for column in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row, column)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    self._process_cell_content(cell, remove_default_words, word_remove_entry)
                
                # Скрытие пустых строк
                if self._should_hide_row(row_data, sheet.max_column):
                    sheet.row_dimensions[row].hidden = True
                
                # Заливка цветом дней недели
                if set_colors and row_data[0].value in self.day_colors:
                    row_data[0].fill = PatternFill(
                        fill_type='solid', 
                        fgColor=self.day_colors[row_data[0].value]
                    )

        new_file_path = file_path.replace('.xlsx', '_updated.xlsx')
        workbook.save(new_file_path)
        workbook.close()
        return new_file_path